# app/routes/scan.py
from flask import Blueprint, redirect, render_template, request, url_for, send_from_directory, flash
from flask_login import login_required
from app.models.website_scanner import WebsiteScanner
from openpyxl import Workbook, load_workbook
from pathlib import Path
from datetime import datetime

bp = Blueprint('scan', __name__)

@bp.route('/scan', methods=["GET", "POST"])
@login_required
def scan():
    if request.method == "GET":
        return redirect(url_for('index.index'))

    form_data = request.form.to_dict()
    values_list = list(set(form_data.values()))

    scanner = WebsiteScanner(values_list)
    scanner.scan_website()

    # Check if an error occurred during scanning
    if scanner.error_occurred:
        # Use flash to send an error message
        flash("Impossible to proceed due to an error.", "error")
        # Redirect back to the same page (the index page in this case)
        return redirect(url_for('index.index'))

    # If no error occurred, proceed as before
    urlResults = scanner.get_urls()
    total_results = len(urlResults)
    total_pages = sum(int(data['page_count']) for data in urlResults)
    average_pages_per_result = total_pages / total_results if total_results > 0 else 0

    return render_template('res.html',
                           urlResults=urlResults,
                           sum_page_count=total_pages,
                           average_pages_per_result=average_pages_per_result)

@bp.route('/scan/generate/', methods=["POST"])
@login_required
def generate():
    data = request.json
    current_dir = Path(__file__).parent
    file_path = current_dir.parent.parent / \
        "resources" / "Desktop_Estimation_CLIENT_SITE_ANNEE.xlsx"

    # Load the workbook:
    wb = load_workbook(file_path)
    ws = wb.active
    start_row = 2
    for item in data:
        ws.cell(row=start_row, column=2, value=item['pageName'])
        ws.cell(row=start_row, column=3, value=item['url'])
        ws.cell(row=start_row, column=4, value=item['components'])
        # ... Populate other columns as needed
        start_row += 1

    # Creating a unique filename with the current date and time
    current_time_str = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    output_filename = f'output_{current_time_str}.xlsx'

    excel_file_path = current_dir.parent.parent / "Dowloads"
    
    # Ensure the directory exists
    excel_file_path.mkdir(parents=True, exist_ok=True)
    
    # Joining paths correctly using the `/` operator
    full_path = excel_file_path / output_filename

    # Save the workbook in the specified directory
    wb.save(full_path)

    # Note: send_from_directory expects string paths, so we convert the directory path to a string
    return send_from_directory(directory=str(excel_file_path), path=output_filename, as_attachment=True)

